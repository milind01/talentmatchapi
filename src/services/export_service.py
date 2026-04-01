"""Shortlist export service - generates Excel and CSV exports."""
import io
import csv
import logging
from typing import List, Dict, Any

logger = logging.getLogger(__name__)


def export_shortlist_excel(candidates: List[Dict[str, Any]], jd_title: str = "Job") -> bytes:
    """
    Generate an Excel (.xlsx) file of shortlisted candidates.
    Returns bytes ready for FastAPI StreamingResponse.
    Requires: pip install openpyxl
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise ImportError("Install openpyxl: pip install openpyxl")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Shortlisted Candidates"

    # Header styling
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    headers = [
        "Rank", "Name", "Email", "Phone", "Current Role",
        "Experience (yrs)", "Education", "Skills",
        "Match Score (%)", "Skills Score", "Exp Score", "Edu Score",
        "Summary"
    ]

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Freeze header row
    ws.freeze_panes = "A2"

    # Data rows
    for row_idx, c in enumerate(candidates, 2):
        breakdown = c.get("match_breakdown") or {}
        skills_list = c.get("skills") or []

        ws.cell(row=row_idx, column=1, value=c.get("rank", row_idx - 1))
        ws.cell(row=row_idx, column=2, value=c.get("candidate_name", "N/A"))
        ws.cell(row=row_idx, column=3, value=c.get("email", "N/A"))
        ws.cell(row=row_idx, column=4, value=c.get("phone", "N/A"))
        ws.cell(row=row_idx, column=5, value=c.get("current_role", "N/A"))
        ws.cell(row=row_idx, column=6, value=c.get("experience_years", "N/A"))
        ws.cell(row=row_idx, column=7, value=c.get("education", "N/A"))
        ws.cell(row=row_idx, column=8, value=", ".join(skills_list[:10]))
        ws.cell(row=row_idx, column=9, value=c.get("match_score", 0))
        ws.cell(row=row_idx, column=10, value=breakdown.get("skills", "N/A"))
        ws.cell(row=row_idx, column=11, value=breakdown.get("experience", "N/A"))
        ws.cell(row=row_idx, column=12, value=breakdown.get("education", "N/A"))
        ws.cell(row=row_idx, column=13, value=c.get("summary", ""))

        # Color code by score
        score = c.get("match_score", 0)
        score_cell = ws.cell(row=row_idx, column=9)
        if score >= 75:
            score_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        elif score >= 50:
            score_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        else:
            score_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    # Auto-fit column widths
    col_widths = [8, 20, 25, 15, 25, 15, 25, 40, 15, 12, 12, 12, 50]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # Add title row above headers (insert at row 1, push headers to row 2)
    ws.insert_rows(1)
    title_cell = ws.cell(row=1, column=1, value=f"Shortlisted Candidates - {jd_title}")
    title_cell.font = Font(bold=True, size=14)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()


def export_shortlist_csv(candidates: List[Dict[str, Any]]) -> str:
    """
    Generate a CSV string of shortlisted candidates.
    """
    output = io.StringIO()
    headers = [
        "Rank", "Name", "Email", "Phone", "Current Role",
        "Experience (yrs)", "Education", "Skills", "Match Score (%)",
        "Skills Score", "Exp Score", "Edu Score", "Summary"
    ]

    writer = csv.DictWriter(output, fieldnames=headers, extrasaction="ignore")
    writer.writeheader()

    for c in candidates:
        breakdown = c.get("match_breakdown") or {}
        skills_list = c.get("skills") or []
        writer.writerow({
            "Rank": c.get("rank", ""),
            "Name": c.get("candidate_name", ""),
            "Email": c.get("email", ""),
            "Phone": c.get("phone", ""),
            "Current Role": c.get("current_role", ""),
            "Experience (yrs)": c.get("experience_years", ""),
            "Education": c.get("education", ""),
            "Skills": ", ".join(skills_list[:10]),
            "Match Score (%)": c.get("match_score", ""),
            "Skills Score": breakdown.get("skills", ""),
            "Exp Score": breakdown.get("experience", ""),
            "Edu Score": breakdown.get("education", ""),
            "Summary": c.get("summary", ""),
        })

    return output.getvalue()
